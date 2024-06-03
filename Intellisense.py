7.2 CODE 
import openpyxl 
import cv2 
import pytesseract 
import numpy as np 
import re 
import smtplib 
from pydrive.auth import GoogleAuth 
from pydrive.drive import GoogleDrive 
from datetime import datetime 
from PIL import Image 
from openpyxl import Workbook 
from openpyxl.styles import NamedStyle 
from reportlab.lib.pagesizes import letter 
from reportlab.pdfgen import canvas 
from reportlab.lib.units import inch 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 
HR26DK8337 = 'phancode23@gmail.com' 
gauth = GoogleAuth() 
drive = GoogleDrive(gauth) 
folder = '13354mIilaGz8vJeSdlBNxQN5OdX92QaH'  
download_directory = "C:/Users/phane/OneDrive/Documents/Inpr1/data base/Vehicle image" 
# Get today's date 
today_date = datetime.today().date() 
# Format the date as "YYYY-MM-DD" 
date = today_date.strftime("%Y-%m-%d") 
# Remove hyphens from the formatted date 
today = date.replace("-", "") 
# Print the formatted date without hyphens 
print("Today's date without hyphens:", today) 
# Specify the name of the subfolder you want to download 
subfolder_name = today 
# Find the subfolder in the main folder 
subfolder_id = None 
file_list = drive.ListFile({'q': f"'{folder}' in parents and trashed=false"}).GetList() 
for file in file_list: 
if file['title'] == subfolder_name and file['mimeType'] == 'application/vnd.google-apps.folder': 
subfolder_id = file['id'] 
break 
# Download files from the subfolder 
if subfolder_id: 
subfolder_file_list = drive.ListFile({'q': f"'{subfolder_id}' in parents and trashed=false"}).GetList() 
for index, file in enumerate(subfolder_file_list): 
print(index + 1, 'file downloaded:', file['title']) 
file.GetContentFile(os.path.join(download_directory, file['title'])) 
# Delete files from the subfolder after downloading (optional) 
# file.Delete() 
else: 
print(f"Subfolder '{subfolder_name}' not found in the main folder.") 
# Set the path to the Tesseract executable (modify this according to your installation) 
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' 
def crop_image(input_path, output_path): 
# Open the image 
img = Image.open(input_path) 
# Get the dimensions of the image 
width, height = img.size 
# Calculate cropping dimensions 
top_crop = int(0.15 * height) 
left_crop = int(0.2 * width) 
right_crop = width - left_crop 
lower_crop = height - top_crop 
# Crop the image 
cropped_img = img.crop((left_crop, top_crop, right_crop, lower_crop)) 
# Save the cropped image 
cropped_img.save(output_path) 
from PIL import Image, ImageFilter
def enlarge_image(input_path, output_path, scale_factor): 
# Open the image file 
image = Image.open(input_path) 
# Get the original width and height 
original_width, original_height = image.size  
# Calculate the new width and height based on the scale factor 
new_width = int(original_width * scale_factor)  
new_height = int(original_height * scale_factor) 
# Resize the image using the BICUBIC resampling filter 
enlarged_image = image.resize((new_width, new_height), Image.BICUBIC) 
# Save the enlarged image 
enlarged_image.save(output_path) 
from PIL import Image, ImageEnhance 
def increase_contrast(input_path, output_path, factor): 
# Open the image file 
image = Image.open(input_path) 
# Create a contrast enhancer 
enhancer = ImageEnhance.Contrast(image) 
# Increase the contrast by the specified factor 
contrast_image = enhancer.enhance(factor) 
# Save the image with increased contrast 
contrast_image.save(output_path) 
def image_operations(directory): 
# Define the file extensions for images you want to include 
image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp'] 
# List all files in the specified directory 
all_files = os.listdir(directory) 
# Filter only the image files 
image_files = [file for file in all_files if any(file.lower().endswith(ext) for ext in image_extensions)] 
# Print the current working directory 
print("Current working directory:", os.getcwd()) 
# Print the paths of the found image files 
if not image_files: 
print(f"No images found in {directory}") 
else:
print("Image files found:") 
for image_path in image_files: 
directory1 = directory.replace("/","\\") 
path = os.path.join(directory1, image_path) 
path1 = path.replace("\\","/") 
crop_image(path1, path1) 
enlarge_image(path1, path1, 2.0) 
increase_contrast(path1,path1, 2.0) 
# Replace 'your_directory_path' with the path to the directory containing the images 
vehicleimage_directory_path = 'C:/Users/phane/OneDrive/Documents/Inpr1/data base/Vehicle image' 
image_operations(vehicleimage_directory_path) 
def process_image(input_path, output_directory): 
# Load the image 
image = cv2.imread(input_path)  
# Preprocess the image for better text extraction 
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) 
gray = cv2.bilateralFilter(gray, 11, 17, 17) 
# Apply edge detection to find the contours of the number plate 
edges = cv2.Canny(gray, 30, 200) 
contours, _ = cv2.findContours(edges, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE) 
contours = sorted(contours, key=cv2.contourArea, reverse=True)[:10] 
# Find the number plate contour based on its shape and dimensions 
number_plate_contour = None 
for contour in contours:  
perimeter = cv2.arcLength(contour, True) 
approx = cv2.approxPolyDP(contour, 0.018 * perimeter, True) 
if len(approx) == 4: 
number_plate_contour = approx 
break 
if number_plate_contour is not None: 
# Extract the number plate region from the image 
mask = np.zeros(gray.shape, np.uint8) 
cv2.drawContours(mask, [number_plate_contour], 0, 255, -1) 
masked_image = cv2.bitwise_and(gray, gray, mask=mask) 
# Save the processed image in the output directory 
   output_path = os.path.join(output_directory, f"processed_{os.path.basename(input_path)}") 
        cv2.imwrite(output_path, masked_image) 
        # Apply OCR (Optical Character Recognition) to recognize the text from the number plate region 
        text = pytesseract.image_to_string(masked_image, config='--psm 11') 
        def check_and_fix_string(input_string): 
            while input_string and not input_string[0].isalpha(): 
                input_string = input_string[1:] 
            while input_string and not input_string[-1].isdigit(): 
                input_string = input_string[:-1] 
            return input_string 
        result = check_and_fix_string(text) 
        result1 = result.replace(" ", "") 
        return result1 
     else: 
        print(f"No number plate contour found in {os.path.basename(input_path)}") 
        # Input and output directories 
       input_directory = "C:/Users/phane/OneDrive/Documents/Inpr1/data base/Vehicle image/" 
       output_directory = "C:/Users/phane/OneDrive/Documents/Inpr1/data base/Number plate images/" 
        # Iterate over all files in the input directory 
        for filename in os.listdir(input_directory): 
        if filename.endswith((".jpg", ".jpeg", ".png")):  # Adjust file extensions as needed 
        image_path = os.path.join(input_directory, filename) 
        process_image(image_path, output_directory) 
       result2 = process_image(image_path, output_directory) 
       def create_or_load_excel(filename): 
       try:  
        workbook = openpyxl.load_workbook(filename)  
       except FileNotFoundError: 
        workbook = Workbook() 
       return workbook 
       def add_row(sheet, time, amount): 
       next_row = sheet.max_row + 1 
       sheet.cell(row=next_row, column=1, value=time)  
       sheet.cell(row=next_row, column=2, value=amount) 
     def set_column_width(sheet, column, width): 
sheet.column_dimensions[column].width = width 
def save_to_excel(workbook, filename):  
workbook.save(filename) 
def convert_to_pdf(excel_filename, pdf_filename, image_filename): 
workbook = openpyxl.load_workbook(excel_filename) 
sheet = workbook.active 
pdf_canvas = canvas.Canvas(pdf_filename, pagesize=letter) 
# Set font and size 
pdf_canvas.setFont("Helvetica", 12) 
# Set the position for field names 
pdf_canvas.drawString(100, 750, "Time") 
pdf_canvas.drawString(250, 750, "Amount Deducted")  # Adjusted X-coordinate for "Amount 
Deducted" 
# Set the position for data rows 
row_height = 20 
y_coordinate = 730 
for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True): 
time, amount = row 
pdf_canvas.drawString(100, y_coordinate, str(time)) 
pdf_canvas.drawString(250, y_coordinate, str(amount))  # Adjusted X-coordinate for data rows 
y_coordinate -= row_height 
# Calculate the total height of text content 
text_height = (sheet.max_row - 1) * row_height 
# Set the position for the image below the text content 
y_coordinate -= text_height + 50  # Place image below text content with additional spacing 
# Get image dimensions 
image_width = 3 * inch 
image_height = 4 * inch  # Adjust image height as needed 
# Draw image 
pdf_canvas.drawImage(image_filename, 100, y_coordinate - image_height, width=image_width, 
height=image_height) 
pdf_canvas.save() 
def upload_to_google_drive(pdf_filename, drive_folder_id): 
gauth = GoogleAuth() 
gauth.LocalWebserverAuth()
drive = GoogleDrive(gauth) 
file_list = drive.ListFile({'q': f"'{drive_folder_id}' in parents and trashed=false"}).GetList() 
existing_file_id = None 
for file in file_list: 
if file['title'] == os.path.basename(pdf_filename): 
existing_file_id = file['id'] 
break 
if existing_file_id: 
# If the file already exists, update its content and metadata 
existing_file = drive.CreateFile({'id': existing_file_id}) 
existing_file.SetContentFile(pdf_filename) 
existing_file.Upload() 
else: 
# If the file doesn't exist, create a new one 
file = drive.CreateFile({'title': os.path.basename(pdf_filename), 'parents': [{'id': drive_folder_id}]}) 
file.SetContentFile(pdf_filename) 
file.Upload() 
def delete_files_in_directory(directory):  
# List all files in the directory 
files = os.listdir(directory) 
# Iterate through each file and delete it 
for file in files: 
# Construct the full path to the file 
file_path = os.path.join(directory, file) 
# Check if it's a file (not a directory) 
if os.path.isfile(file_path):  
# Delete the file 
os.remove(file_path) 
sheet_name = result2 + ".xlsx" 
default_path = r'C:\Users\phane\OneDrive\Documents\Inpr1\data base\Bills' 
image_filename = r"C:\Users\phane\OneDrive\Documents\Inpr1\data base\Bills\upi.png"  # Provide the 
path to your image file here 
print (f"Extracted number plate number: {result2}") 
workbook = create_or_load_excel(os.path.join(default_path, sheet_name)) 
sheet = workbook.active
if sheet["A1"].value is None or sheet["B1"].value is None: 
sheet["A1"] = "Time" 
sheet["B1"] = "Amount Deducted" 
current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S") 
add_row(sheet, current_time, 170) 
set_column_width(sheet, 'A', len(current_time) + 2) 
excel_filename = os.path.join(default_path, sheet_name) 
pdf_filename = os.path.join(default_path, os.path.splitext(sheet_name)[0] + ".pdf") 
drive_folder_id = '1-iP3vmUwWUjJKz8vB88Uq7ox_yzmodtU' 
save_to_excel(workbook, excel_filename) 
convert_to_pdf(excel_filename, pdf_filename, image_filename) 
upload_to_google_drive(pdf_filename, drive_folder_id) 
directory_path = r"C:\Users\phane\OneDrive\Documents\Inpr1\data base\Vehicle image" 
delete_files_in_directory(directory_path) 
# Access the value of the variable using locals() or globals() 
accounts = locals().get(result2) 
def send_email(sender_email, sender_password, recipient_email, subject, body, attachment_path): 
# Set up SMTP server 
smtp_server = 'smtp.gmail.com'  # Update with your SMTP server 
smtp_port = 587  # Update with your SMTP port 
server = smtplib.SMTP(smtp_server, smtp_port) 
server.starttls() 
server.login(sender_email, sender_password) 
# Create a multipart message 
msg = MIMEMultipart() 
msg['From'] = sender_email 
msg['To'] = recipient_email 
msg['Subject'] = subject 
# Attach body 
msg.attach(MIMEText(body, 'plain')) 
# Attach file 
with open(attachment_path, 'rb') as attachment: 
part = MIMEBase('application', 'octet-stream') 
part.set_payload(attachment.read()) 
encoders.encode_base64(part)
part.add_header( 
'Content-Disposition', 
f'attachment; filename= {attachment_path.split("/")[-1]}') 
msg.attach(part)  
# Send the email 
server.send_message(msg)  
del msg 
server.quit() 
# Example usage 
sender_email = 'phaneendra2k3@gmail.com' 
sender_password = 'ejbk anhp rcvh gczd' 
recipient_email = accounts 
subject = 'Toll Bill' 
body = 'This is the mail sent by National Toll Agency (NTA), please find the attached toll bill below.' 
attachment_path = r"C:\Users\phane\OneDrive\Documents\Inpr1\data base\Bills\HR26DK8337.pdf" 
send_email(sender_email, sender_password, recipient_email, subject, body, attachment_path)
